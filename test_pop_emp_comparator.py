# -*- coding: utf-8 -*-
"""
    Module containing the test function for running the PopEmpComparator without
    running the whole EFS process.
"""

##### IMPORTS #####
# Standard imports
import os
from pathlib import Path
from typing import Dict

# Third party imports
import pytest
import numpy as np
import pandas as pd
from openpyxl import Workbook

# Local imports
from pop_emp_comparator import PopEmpComparator, _excel_column_format
from demand_utilities import utils as du
from external_forecast_system import ExternalForecastSystem


##### CONSTANTS #####
# Constants for running the test functions
MODEL_NAME = "norms_2015"
ITER_NUM = 1
IMPORT_LOC = "Y:/"
EXPORT_LOC = "Y:/"
BASE_YEAR = "2018"
YEARS = [BASE_YEAR, "2020"]
ZONE_COL = "msoa_zone_id"
TEST_MSOAS = ["E02000001", "E02000002", "E02000003", "E02000004", "E02000005"]

# Test DataFrames
BASE_DATA = pd.DataFrame(
    {ZONE_COL: TEST_MSOAS, BASE_YEAR: np.random.randint(1, 10000, 5)}
)
GROWTH_DATA = pd.DataFrame(
    {ZONE_COL: TEST_MSOAS, YEARS[0]: np.random.rand(5), YEARS[1]: np.random.rand(5)}
)
CONSTRAINT_DATA = pd.DataFrame(
    {
        ZONE_COL: TEST_MSOAS,
        YEARS[0]: np.random.randint(1, 10000, 5),
        YEARS[1]: np.random.randint(1, 10000, 5),
    }
)
OUTPUT_POPULATION = pd.DataFrame(
    {
        ZONE_COL: TEST_MSOAS,
        "area_type": [1, 1, 1, 1, 2],
        "traveller_type": [1, 1, 2, 2, 1],
        "soc": [1, 2, 1, 3, 1],
        "ns": [1, 2, 2, 1, 2],
        "ca": [0, 1, 0, 1, 0],
        YEARS[0]: np.random.randint(1, 10000, 5),
        YEARS[1]: np.random.randint(1, 10000, 5),
    }
)
OUTPUT_EMPLOYMENT = pd.DataFrame(
    {
        ZONE_COL: TEST_MSOAS,
        "employment_cat": ["E01", "E01", "E02", "E03", "E02"],
        YEARS[0]: np.random.randint(1, 10000, 5),
        YEARS[1]: np.random.randint(1, 10000, 5),
    }
)
SECTOR_LOOKUP = pd.DataFrame(
    {
        ZONE_COL: TEST_MSOAS,
        "overlap_msoa_split_factor": [0.1, 0.2, 0.7, 0.4, 0.6],
        "tfn_sectors_zone_id": [1, 1, 1, 2, 2],
    }
)


##### FUNCTIONS #####
@pytest.mark.skip(reason="test on real data which can be run by running this module")
def test_real_data():
    """Tests the PopEmpComparator class on data from previous run of EFS. """
    imports, exports, _ = du.build_io_paths(
        IMPORT_LOC,
        EXPORT_LOC,
        MODEL_NAME,
        f"iter{ITER_NUM}",
        ExternalForecastSystem.__version__,
        ExternalForecastSystem._out_dir,
    )
    # Population csv files, locations from ExternalForecastSysten.__init__ parameters
    population_value_file = "population/base_population_2018.csv"
    population_growth_file = "population/future_population_growth.csv"
    population_constraint_file = "population/future_population_values.csv"
    # Employment csv files
    worker_value_file = "employment/base_workers_2018.csv"
    worker_growth_file = "employment/future_workers_growth.csv"
    worker_constraint_file = "employment/future_workers_growth_values.csv"

    # Compare the population inputs and outputs
    pop_comp = PopEmpComparator(
        os.path.join(imports["default_inputs"], population_value_file),
        os.path.join(imports["default_inputs"], population_growth_file),
        os.path.join(imports["default_inputs"], population_constraint_file),
        os.path.join(exports["productions"], "MSOA_population.csv"),
        "population",
        BASE_YEAR,
        sector_grouping_file=os.path.join(
            imports["zoning"], "tfn_sector_msoa_pop_weighted_lookup.csv"
        ),
    )
    pop_comp.write_comparisons(exports["reports"], output_as="csv", year_col=True)
    pop_comp.write_comparisons(exports["reports"], output_as="excel", year_col=True)
    # Compare the employment inputs and outputs
    emp_comp = PopEmpComparator(
        os.path.join(imports["default_inputs"], worker_value_file),
        os.path.join(imports["default_inputs"], worker_growth_file),
        os.path.join(imports["default_inputs"], worker_constraint_file),
        os.path.join(exports["attractions"], "MSOA_employment.csv"),
        "employment",
        BASE_YEAR,
        sector_grouping_file=os.path.join(
            imports["zoning"], "tfn_sector_msoa_emp_weighted_lookup.csv"
        ),
    )
    emp_comp.write_comparisons(exports["reports"], output_as="csv", year_col=True)
    emp_comp.write_comparisons(exports["reports"], output_as="excel", year_col=True)
    return


@pytest.mark.parametrize("style", ["Percent", "Comma [0]", None, "Normal"])
@pytest.mark.parametrize("ignore_rows", [0, 1])
def test_excel_column_format(style: str, ignore_rows: int):
    """Test the _excel_column_format function with different parameters.

    Parameters
    ----------
    style : str
        The style to convert column format to, passed to _excel_column_format.
    ignore_rows : int
        The number of rows to ignore, passed to _excel_column_format.
    """
    wb = Workbook()
    ws = wb.active
    rows = 2

    # Fill some cells
    for r in range(rows):
        ws.cell(row=r + 1, column=1, value=np.random.randint(100))

    _excel_column_format(ws, [style], ignore_rows=ignore_rows)

    style = "Normal" if style is None else style
    # Check styles
    for r in range(rows):
        this_style = "Normal" if r < ignore_rows else style
        new_style = ws.cell(row=r + 1, column=1).style
        assert new_style == this_style, f"'{new_style}' != '{this_style}' for row {r}"


@pytest.fixture(name="test_files")
def fixture_test_files(tmpdir) -> Dict[str, Path]:
    """Write test data to CSVs in temporary directory.

    Parameters
    ----------
    tmpdir : pathlib.Path
        Temporary directory provided by pytest

    Returns
    -------
    Dict[str, Path]
        Paths to the test files.
    """
    paths = {
        "base": tmpdir / "base_data.csv",
        "growth": tmpdir / "growth_data.csv",
        "constraint": tmpdir / "constraint_data.csv",
        "sector_lookup": tmpdir / "sector_lookup.csv",
        "output_population": tmpdir / "output_population.csv",
        "output_employment": tmpdir / "output_employment.csv",
    }
    data = [
        BASE_DATA,
        GROWTH_DATA,
        CONSTRAINT_DATA,
        SECTOR_LOOKUP,
        OUTPUT_POPULATION,
        OUTPUT_EMPLOYMENT,
    ]
    for p, df in zip(paths.values(), data):
        df.to_csv(p, index=False)
    return paths


@pytest.fixture(name="initialise_class")
def fixture_initialise_class(test_files: pytest.fixture) -> Dict[str, PopEmpComparator]:
    """Initialise the PopEmpComarator class for both "population" and "employment".

    Parameters
    ----------
    test_files : pytest.fixture
        Paths to the test files.

    Returns
    -------
    Dict[str, PopEmpComparator]
        Two instances of PopEmpComparator class with the keys "population"
        and "employment".
    """
    comparators = {}
    for i in ("population", "employment"):
        comparators[i] = PopEmpComparator(
            test_files["base"],
            test_files["growth"],
            test_files["constraint"],
            test_files[f"output_{i}"],
            i,
            BASE_YEAR,
            sector_grouping_file=test_files["sector_lookup"],
        )
    return comparators


@pytest.mark.parametrize(
    "data_type,output",
    [
        ("population", OUTPUT_POPULATION),
        ("employment", OUTPUT_EMPLOYMENT),
    ],
)
def test_initialisation(
    initialise_class: pytest.fixture, data_type: str, output: pd.DataFrame
):
    """Test that the PopEmpComparator class initilises variables correctly.

    Parameters
    ----------
    initialise_class : pytest.fixture
        Dictionary containing both initialised classes.
    data_type : str
        Whether to test the 'population' or 'employment' comparisons.
    output : pd.DataFrame
        The output 'population' or 'employment' test data to compare against.
    """
    instance = initialise_class[data_type]
    pd.testing.assert_frame_equal(instance.input_data, BASE_DATA, check_dtype=False)
    pd.testing.assert_frame_equal(
        instance.constraint_data, CONSTRAINT_DATA, check_dtype=False
    )
    pd.testing.assert_frame_equal(instance.output, output, check_dtype=False)

    # Normalise growth data before comparison
    normalised = GROWTH_DATA.copy()
    for y in YEARS:
        normalised[y] = normalised[y] / GROWTH_DATA[BASE_YEAR]
    pd.testing.assert_frame_equal(instance.growth_data, normalised, check_dtype=False)


##### MAIN #####
if __name__ == "__main__":
    test_real_data()
